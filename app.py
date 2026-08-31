import streamlit as st
import pandas as pd
import numpy as np
import io
import plotly.express as px
import plotly.graph_objects as go
from supabase import create_client, Client

# --- BIBLIOTECAS PARA GERAÇÃO DO EXCEL COM GRÁFICOS ---
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

# --- BIBLIOTECAS PARA GERAÇÃO DE PDF ---
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import matplotlib.pyplot as plt

# ==========================================
# CONFIGURAÇÃO DA PÁGINA STREAMLIT
# ==========================================
st.set_page_config(
    page_title="Sistema de Sondagem Rotativa",
    page_icon="⛏️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==========================================
# CONEXÃO COM O SUPABASE
# ==========================================
@st.cache_resource
def init_supabase() -> Client:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

try:
    supabase = init_supabase()
except Exception as e:
    st.error(f"Erro ao conectar ao Supabase: {e}")

# ==========================================
# FUNÇÃO PARA GERAR PLANILHA EXCEL COM GRÁFICOS
# ==========================================
def gerar_excel_formatado(df_manobras: pd.DataFrame, df_paradas: pd.DataFrame, id_furo: str) -> io.BytesIO:
    wb = Workbook()
    
    # --- ABA 1: BOLETIM DE MANOBRAS E GRÁFICO DE PERFIL GEOTÉCNICO ---
    ws = wb.active
    ws.title = f"Furo {id_furo}"
    ws.views.sheetView[0].showGridLines = True
    
    # Estilos Visuais
    cor_cabecalho = "1E293B"
    cor_linha_par = "F8FAFC"
    
    fonte_titulo = Font(name="Calibri", size=14, bold=True, color="1E293B")
    fonte_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fonte_dados = Font(name="Calibri", size=10)
    
    preenchimento_header = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type="solid")
    preenchimento_zebrado = PatternFill(start_color=cor_linha_par, end_color=cor_linha_par, fill_type="solid")
    
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center")
    
    borda_fina = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Cabeçalho Principal
    ws.append([f"BOLETIM DIÁRIO DE SONDAGEM ROTATIVA - FURO: {id_furo}"])
    ws.cell(row=1, column=1).font = fonte_titulo
    ws.append([]) # Linha em branco
    
    # Colunas do Boletim
    colunas = [
        "De (m)", "Até (m)", "Avanço (m)", "Recup. (m)", 
        "Recup. (%)", "RQD (m)", "RQD (%)", "Qualidade RQD", 
        "Litologia", "Caixa Nº", "Retorno Fluido"
    ]
    ws.append(colunas)
    
    linha_header = 3
    for col_idx in range(1, len(colunas) + 1):
        cell = ws.cell(row=linha_header, column=col_idx)
        cell.font = fonte_header
        cell.fill = preenchimento_header
        cell.alignment = alinhamento_centro
        cell.border = borda_fina
    
    # Preenchimento das Linhas de Manobra
    dados_cols = [
        "de_m", "ate_m", "avanco_m", "recup_m", 
        "recup_pct", "rqd_m", "rqd_pct", "qualidade_rqd", 
        "litologia", "caixa_num", "perda_agua"
    ]
    
    for r_idx, row in df_manobras[dados_cols].iterrows():
        row_num = linha_header + 1 + r_idx
        val_lista = row.tolist()
        ws.append(val_lista)
        
        is_par = (r_idx % 2 == 0)
        for c_idx in range(1, len(val_lista) + 1):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.font = fonte_dados
            cell.border = borda_fina
            
            if is_par:
                cell.fill = preenchimento_zebrado
                
            # Máscaras de Formatação
            if c_idx in [1, 2, 3, 4, 6]:
                cell.number_format = '0.00" m"'
                cell.alignment = alinhamento_centro
            elif c_idx in [5, 7]:
                cell.number_format = '0.0"%"'
                cell.alignment = alinhamento_centro
            elif c_idx in [8, 10, 11]:
                cell.alignment = alinhamento_centro
            else:
                cell.alignment = alinhamento_esquerda

    # Ajuste automático de largura das colunas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= linha_header:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 📈 Adição do Gráfico de Linha Geotécnico (Recuperação vs RQD)
    if not df_manobras.empty:
        chart_rqd = LineChart()
        chart_rqd.title = "Perfil Geotécnico: Recuperação vs RQD (%)"
        chart_rqd.style = 13
        chart_rqd.y_axis.title = "Percentual (%)"
        chart_rqd.x_axis.title = "Profundidade (m)"
        chart_rqd.width = 16
        chart_rqd.height = 10
        
        # Referência de dados: Colunas 5 (Recup %) e 7 (RQD %)
        dados_rqd = Reference(ws, min_col=5, min_row=3, max_col=7, max_row=ws.max_row)
        categorias = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
        
        chart_rqd.add_data(dados_rqd, titles_from_data=True)
        chart_rqd.set_categories(categorias)
        ws.add_chart(chart_rqd, "M3")

    # --- ABA 2: REGISTRO DE DOWNTIME E GRÁFICO DE PARADAS ---
    if not df_paradas.empty:
        ws_paradas = wb.create_sheet(title="Downtime & Paradas")
        ws_paradas.views.sheetView[0].showGridLines = True
        
        ws_paradas.append(["RELATÓRIO DE PARADAS E TEMPOS IMOBILIZADOS"])
        ws_paradas.cell(row=1, column=1).font = fonte_titulo
        ws_paradas.append([])
        
        col_p = ["Categoria / Motivo", "Duração (Horas)", "Observação", "Data/Hora"]
        ws_paradas.append(col_p)
        
        for col_idx in range(1, len(col_p) + 1):
            cell = ws_paradas.cell(row=3, column=col_idx)
            cell.font = fonte_header
            cell.fill = preenchimento_header
            cell.alignment = alinhamento_centro
            cell.border = borda_fina
            
        for r_idx, row in df_paradas[["categoria", "horas", "observacao", "data_hora"]].iterrows():
            row_num = 4 + r_idx
            val_lista = row.tolist()
            ws_paradas.append(val_lista)
            
            for c_idx in range(1, len(val_lista) + 1):
                cell = ws_paradas.cell(row=row_num, column=c_idx)
                cell.font = fonte_dados
                cell.border = borda_fina
                if c_idx == 2:
                    cell.number_format = '0.0" h"'
                    cell.alignment = alinhamento_centro
                else:
                    cell.alignment = alinhamento_esquerda
                    
        # 📊 Adição do Gráfico de Barras de Downtime
        chart_bar = BarChart()
        chart_bar.type = "col"
        chart_bar.style = 10
        chart_bar.title = "Horas Paradas por Categoria"
        chart_bar.y_axis.title = "Horas"
        chart_bar.x_axis.title = "Motivo"
        chart_bar.width = 15
        chart_bar.height = 9
        
        dados_bar = Reference(ws_paradas, min_col=2, min_row=3, max_row=ws_paradas.max_row)
        cats_bar = Reference(ws_paradas, min_col=1, min_row=4, max_row=ws_paradas.max_row)
        
        chart_bar.add_data(dados_bar, titles_from_data=True)
        chart_bar.set_categories(cats_bar)
        ws_paradas.add_chart(chart_bar, "F3")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# FUNÇÃO PARA GERAR RELATÓRIO EM PDF
# ==========================================
def gerar_pdf_relatorio(df_manobras: pd.DataFrame, id_furo: str) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, leading=20, textColor=colors.HexColor('#1E293B'), alignment=1)
    
    story.append(Paragraph(f"<b>BOLETIM DE SONDAGEM ROTATIVA - FURO {id_furo}</b>", title_style))
    story.append(Spacer(1, 15))
    
    # Gerar Gráfico estático via Matplotlib
    fig, ax = plt.subplots(figsize=(6, 2.5))
    ax.plot(df_manobras['ate_m'], df_manobras['recup_pct'], label='Recuperação (%)', color='#2563EB', marker='o')
    ax.plot(df_manobras['ate_m'], df_manobras['rqd_pct'], label='RQD (%)', color='#059669', marker='s')
    ax.set_ylabel('Percentual (%)')
    ax.set_xlabel('Profundidade (m)')
    ax.set_ylim(0, 105)
    ax.grid(True, linestyle='--', alpha=0.5)
    ax.legend(loc='lower right')
    
    img_buf = io.BytesIO()
    plt.savefig(img_buf, format='png', bbox_inches='tight', dpi=150)
    img_buf.seek(0)
    plt.close(fig)
    
    story.append(Image(img_buf, width=480, height=200))
    story.append(Spacer(1, 15))
    
    # Tabela de Dados
    table_data = [["De", "Até", "Avanço", "Recup %", "RQD %", "Qualidade", "Litologia"]]
    for _, row in df_manobras.iterrows():
        table_data.append([
            f"{row['de_m']:.2f}", f"{row['ate_m']:.2f}", f"{row['avanco_m']:.2f}",
            f"{row['recup_pct']:.1f}%", f"{row['rqd_pct']:.1f}%", row['qualidade_rqd'], row['litologia']
        ])
        
    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))
    
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer

# ==========================================
# INTERFACE PRINCIPAL - STREAMLIT
# ==========================================
st.title("⛏️ Gestão de Sondagem Rotativa & Boletins")

tabs = st.tabs(["📝 Lançamento de Manobras", "⏱️ Registro de Downtime", "📊 Dashboard Geotécnico", "📥 Exportação e Relatórios"])

# ------------------------------------------
# ABA 1: LANÇAMENTO DE MANOBRAS
# ------------------------------------------
with tabs[0]:
    st.header("Novo Lançamento de Manobra")
    
    with st.form("form_manobra", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            furo = st.text_input("Identificador do Furo", value="FD-01")
            de_m = st.number_input("De (m)", min_value=0.0, step=0.1, format="%.2f")
            ate_m = st.number_input("Até (m)", min_value=0.0, step=0.1, format="%.2f")
        with col2:
            recup_m = st.number_input("Recuperação (m)", min_value=0.0, step=0.1, format="%.2f")
            rqd_m = st.number_input("RQD (m)", min_value=0.0, step=0.1, format="%.2f")
            caixa_num = st.number_input("Número da Caixa", min_value=1, step=1)
        with col3:
            litologia = st.selectbox("Litologia Domina", ["Solo de Alteração", "Basalto Alterado", "Basalto Sano", "Diabásio", "Arenito"])
            perda_agua = st.selectbox("Retorno de Fluido", ["Total (100%)", "Médio (50-80%)", "Baixo (10-40%)", "Perda Total (0%)"])
            
        submitted = st.form_submit_button("💾 Salvar Manobra no Supabase")
        
        if submitted:
            avanco = ate_m - de_m
            if avanco <= 0:
                st.error("O valor 'Até (m)' deve ser maior que 'De (m)'.")
            else:
                recup_pct = (recup_m / avanco) * 100 if avanco > 0 else 0
                rqd_pct = (rqd_m / avanco) * 100 if avanco > 0 else 0
                
                # Classificação de Qualidade RQD
                if rqd_pct < 25: qualidade = "Muito Má"
                elif rqd_pct < 50: qualidade = "Má"
                elif rqd_pct < 75: qualidade = "Regular"
                elif rqd_pct < 90: qualidade = "Boa"
                else: qualidade = "Excelente"
                
                payload = {
                    "furo": furo, "de_m": de_m, "ate_m": ate_m, "avanco_m": avanco,
                    "recup_m": recup_m, "recup_pct": recup_pct, "rqd_m": rqd_m,
                    "rqd_pct": rqd_pct, "qualidade_rqd": qualidade, "litologia": litologia,
                    "caixa_num": caixa_num, "perda_agua": perda_agua
                }
                
                try:
                    supabase.table("manobras").insert(payload).execute()
                    st.success("Manobra cadastrada com sucesso!")
                except Exception as ex:
                    st.error(f"Falha ao gravar registro: {ex}")

# ------------------------------------------
# ABA 2: REGISTRO DE DOWNTIME
# ------------------------------------------
with tabs[1]:
    st.header("Registro de Paradas da Sonda")
    
    with st.form("form_downtime", clear_on_submit=True):
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            furo_dt = st.text_input("Furo Associado", value="FD-01")
            categoria = st.selectbox("Motivo da Parada", [
                "Manutenção Preventiva", "Manutenção Corretiva", 
                "Aguardando Mudança de Praça", "Abastecimento / Água", 
                "Condições Climáticas (Chuva)", "Mudança de Turno"
            ])
        with col_d2:
            horas = st.number_input("Duração (Horas)", min_value=0.1, max_value=24.0, step=0.5)
            obs = st.text_area("Observações Operacionais")
            
        sub_dt = st.form_submit_button("🛑 Registrar Parada")
        if sub_dt:
            payload_dt = {"furo": furo_dt, "categoria": categoria, "horas": horas, "observacao": obs}
            try:
                supabase.table("downtime").insert(payload_dt).execute()
                st.success("Parada registrada com sucesso!")
            except Exception as ex:
                st.error(f"Erro ao salvar downtime: {ex}")

# ------------------------------------------
# CONSULTA DE DADOS PARA EXIBIÇÃO
# ------------------------------------------
try:
    res_manobras = supabase.table("manobras").select("*").execute()
    df_manobras_full = pd.DataFrame(res_manobras.data)
    
    res_paradas = supabase.table("downtime").select("*").execute()
    df_paradas_full = pd.DataFrame(res_paradas.data)
except Exception:
    df_manobras_full = pd.DataFrame()
    df_paradas_full = pd.DataFrame()

# ------------------------------------------
# ABA 3: DASHBOARD GEOTÉCNICO
# ------------------------------------------
with tabs[2]:
    st.header("Visualização Interativa de Perfil de Sondagem")
    
    if not df_manobras_full.empty:
        furo_sel = st.selectbox("Selecione o Furo:", df_manobras_full['furo'].unique())
        df_furo = df_manobras_full[df_manobras_full['furo'] == furo_sel].sort_values("de_m")
        
        m1, m2, m3 = st.columns(3)
        m1.metric("Avanço Total", f"{df_furo['ate_m'].max():.2f} m")
        m2.metric("Média de Recuperação", f"{df_furo['recup_pct'].mean():.1f}%")
        m3.metric("Média RQD", f"{df_furo['rqd_pct'].mean():.1f}%")
        
        # Gráfico Interativo com Plotly
        fig_plotly = go.Figure()
        fig_plotly.add_trace(go.Scatter(x=df_furo['ate_m'], y=df_furo['recup_pct'], mode='lines+markers', name='Recuperação (%)'))
        fig_plotly.add_trace(go.Scatter(x=df_furo['ate_m'], y=df_furo['rqd_pct'], mode='lines+markers', name='RQD (%)'))
        fig_plotly.update_layout(title=f"Perfil Geotécnico do Furo {furo_sel}", xaxis_title="Profundidade (m)", yaxis_title="Percentual (%)")
        st.plotly_chart(fig_plotly, use_container_width=True)
    else:
        st.info("Nenhum dado de manobra cadastrado ainda.")

# ------------------------------------------
# ABA 4: EXPORTAÇÃO DE RELATÓRIOS (EXCEL E PDF)
# ------------------------------------------
with tabs[3]:
    st.header("Central de Downloads e Exportação")
    
    if not df_manobras_full.empty:
        furo_exp = st.selectbox("Selecione o Furo para Exportar:", df_manobras_full['furo'].unique(), key="exp_furo")
        df_manobras_exp = df_manobras_full[df_manobras_full['furo'] == furo_exp].sort_values("de_m")
        
        if not df_paradas_full.empty:
            df_paradas_exp = df_paradas_full[df_paradas_full['furo'] == furo_exp]
        else:
            df_paradas_exp = pd.DataFrame(columns=["categoria", "horas", "observacao", "data_hora"])
            
        c1, c2 = st.columns(2)
        
        with c1:
            st.subheader("Planilha Excel (.XLSX)")
            st.write("Ficheiro formatado com duas abas e **gráficos do Excel inseridos automaticamente**.")
            
            excel_data = gerar_excel_formatado(df_manobras_exp, df_paradas_exp, furo_exp)
            st.download_button(
                label="📊 Baixar Planilha Excel com Gráficos",
                data=excel_data,
                file_name=f"Boletim_Sondagem_{furo_exp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            
        with c2:
            st.subheader("Relatório PDF (.PDF)")
            st.write("Documento impresso pronto contendo a tabela formatada e o gráfico estático.")
            
            pdf_data = gerar_pdf_relatorio(df_manobras_exp, furo_exp)
            st.download_button(
                label="📄 Baixar Relatório PDF Formatado",
                data=pdf_data,
                file_name=f"Relatorio_Sondagem_{furo_exp}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
    else:
        st.warning("Cadastre dados na aba de manobras para liberar as opções de exportação.")
